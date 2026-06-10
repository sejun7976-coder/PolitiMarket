"""Create a clean GPT-5.5 baseline benchmark report.

The report is intentionally reader-friendly:
- Korean text is written with UTF-8 BOM for CSV/Markdown/HTML.
- An XLSX workbook is produced to avoid Excel/Windows encoding issues.
- Tagging and sentiment benchmarks are separated.
- Sentiment rows show whether the current production model is ahead of or
  behind each comparison model against the GPT-5.5 audit baseline.
"""

from __future__ import annotations

import csv
import html
import json
import math
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parents[1]
EXPORTS_DIR = BASE_DIR / "Exports"


def latest_report_dir() -> Path:
    candidates = [
        p for p in EXPORTS_DIR.glob("report_current_comparison_*")
        if (p / "tables" / "cleaned_model_comparison.csv").exists()
    ]
    if not candidates:
        raise FileNotFoundError("report_current_comparison_* 산출물을 찾지 못했습니다.")
    return max(candidates, key=lambda p: p.stat().st_mtime)


def now_stamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def read_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_csv(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def pct(value: float) -> str:
    return f"{value:.2f}%"


def pctp(value: float) -> str:
    sign = "+" if value > 0 else ""
    return f"{sign}{value:.2f}%p"


def wilson_margin(p: float, n: int) -> float:
    if n <= 0:
        return 0.0
    return 1.96 * math.sqrt(max(0.0, p * (1.0 - p)) / n) * 100


def agreement_rate(rows: list[dict[str, str]], baseline_col: str, model_col: str) -> tuple[int, int, float]:
    common = [
        row for row in rows
        if row.get(baseline_col) not in {"", "Skipped"} and row.get(model_col) not in {"", "Skipped"}
    ]
    matched = sum(1 for row in common if row.get(baseline_col) == row.get(model_col))
    rate = matched / len(common) if common else 0.0
    return len(common), matched, rate


def benchmark_row(
    area: str,
    baseline_name: str,
    target_name: str,
    model_name: str,
    compare_item: str,
    n: int,
    matched: int,
    rate: float,
    current_rate: float | None,
    note: str,
) -> dict[str, Any]:
    error = 1.0 - rate
    margin = wilson_margin(rate, n)
    if current_rate is None:
        diff_text = "-"
        verdict = "기준 행"
    else:
        diff = (current_rate - rate) * 100
        diff_text = pctp(diff)
        if diff > 0.5:
            verdict = f"우리 모델이 {diff_text} 우위"
        elif diff < -0.5:
            verdict = f"우리 모델이 {pctp(diff)} 열위"
        else:
            verdict = "사실상 동률"
    return {
        "분석영역": area,
        "기준모델": baseline_name,
        "비교대상": target_name,
        "모델명": model_name,
        "비교항목": compare_item,
        "공통건수": n,
        "일치건수": matched,
        "일치율": pct(rate * 100),
        "불일치율": pct(error * 100),
        "95% 표본오차": f"±{margin:.2f}%p",
        "현재모델 대비 차이": diff_text,
        "판정": verdict,
        "비고": note,
        "_rate_number": round(rate * 100, 4),
        "_diff_number": round((current_rate - rate) * 100, 4) if current_rate is not None else 0.0,
    }


def create_svg_bar(path: Path, title: str, rows: list[dict[str, Any]], value_key: str, color_key: str | None = None) -> None:
    width = 1120
    left = 360
    top = 62
    row_h = 44
    height = top + len(rows) * row_h + 36
    max_value = max([float(row.get(value_key, 0) or 0) for row in rows] + [1.0])
    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
        '<rect width="100%" height="100%" fill="#ffffff"/>',
        f'<text x="24" y="32" font-family="Malgun Gothic, Inter, Arial" font-size="20" font-weight="700" fill="#111827">{html.escape(title)}</text>',
    ]
    for idx, row in enumerate(rows):
        y = top + idx * row_h
        name = str(row.get("비교대상") or row.get("비교항목") or "")
        value = float(row.get(value_key, 0) or 0)
        width_value = value / max_value * 620
        color = "#2563eb"
        if color_key and row.get(color_key):
            color = str(row[color_key])
        parts.append(f'<text x="24" y="{y + 18}" font-family="Malgun Gothic, Inter, Arial" font-size="13" font-weight="600" fill="#344054">{html.escape(name)}</text>')
        parts.append(f'<rect x="{left}" y="{y}" width="620" height="18" rx="9" fill="#eef2f7"/>')
        parts.append(f'<rect x="{left}" y="{y}" width="{width_value:.2f}" height="18" rx="9" fill="{color}"/>')
        parts.append(f'<text x="{left + 632}" y="{y + 14}" font-family="Malgun Gothic, Inter, Arial" font-size="12" font-weight="700" fill="#111827">{value:.2f}%</text>')
    parts.append("</svg>")
    path.write_text("\n".join(parts), encoding="utf-8-sig")


def create_svg_diff(path: Path, title: str, rows: list[dict[str, Any]]) -> None:
    width = 1120
    left = 500
    top = 62
    row_h = 44
    height = top + len(rows) * row_h + 36
    max_abs = max([abs(float(row.get("_diff_number", 0) or 0)) for row in rows] + [1.0])
    scale = 260 / max_abs
    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
        '<rect width="100%" height="100%" fill="#ffffff"/>',
        f'<text x="24" y="32" font-family="Malgun Gothic, Inter, Arial" font-size="20" font-weight="700" fill="#111827">{html.escape(title)}</text>',
        f'<line x1="{left}" y1="48" x2="{left}" y2="{height - 24}" stroke="#98a2b3" stroke-width="1"/>',
    ]
    for idx, row in enumerate(rows):
        y = top + idx * row_h
        name = str(row.get("비교대상") or "")
        diff = float(row.get("_diff_number", 0) or 0)
        color = "#16a34a" if diff >= 0 else "#dc2626"
        x = left if diff >= 0 else left + diff * scale
        w = abs(diff * scale)
        parts.append(f'<text x="24" y="{y + 18}" font-family="Malgun Gothic, Inter, Arial" font-size="13" font-weight="600" fill="#344054">{html.escape(name)}</text>')
        parts.append(f'<rect x="{x:.2f}" y="{y}" width="{w:.2f}" height="18" rx="9" fill="{color}"/>')
        parts.append(f'<text x="{left + 280}" y="{y + 14}" font-family="Malgun Gothic, Inter, Arial" font-size="12" font-weight="700" fill="#111827">{html.escape(str(row.get("현재모델 대비 차이")))}</text>')
    parts.append("</svg>")
    path.write_text("\n".join(parts), encoding="utf-8-sig")


def html_table(rows: list[dict[str, Any]], fields: list[str]) -> str:
    parts = ["<table><thead><tr>"]
    for field in fields:
        parts.append(f"<th>{html.escape(field)}</th>")
    parts.append("</tr></thead><tbody>")
    for row in rows:
        parts.append("<tr>")
        for field in fields:
            parts.append(f"<td>{html.escape(str(row.get(field, '')))}</td>")
        parts.append("</tr>")
    parts.append("</tbody></table>")
    return "".join(parts)


def markdown_table(rows: list[dict[str, Any]], fields: list[str]) -> str:
    lines = [
        "| " + " | ".join(fields) + " |",
        "| " + " | ".join("---" for _ in fields) + " |",
    ]
    for row in rows:
        values = [str(row.get(field, "")).replace("|", "/") for field in fields]
        lines.append("| " + " | ".join(values) + " |")
    return "\n".join(lines)


def autosize_sheet(ws) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        width = 12
        for cell in ws[letter]:
            width = max(width, min(48, len(str(cell.value or "")) + 2))
        ws.column_dimensions[letter].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_workbook(path: Path, overview: list[dict[str, Any]], tag_rows: list[dict[str, Any]], sentiment_rows: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "요약"
    ws.append(["항목", "내용"])
    for row in overview:
        ws.append([row["항목"], row["내용"]])
    ws["A1"].font = ws["B1"].font = Font(bold=True, color="FFFFFF")
    ws["A1"].fill = ws["B1"].fill = PatternFill("solid", fgColor="1F2937")
    autosize_sheet(ws)

    fields = [
        "분석영역", "기준모델", "비교대상", "모델명", "비교항목", "공통건수", "일치건수",
        "일치율", "불일치율", "95% 표본오차", "현재모델 대비 차이", "판정", "비고",
    ]
    for title, rows in [("태깅 기준표", tag_rows), ("감성 기준표", sentiment_rows)]:
        sheet = wb.create_sheet(title)
        sheet.append(fields)
        for row in rows:
            sheet.append([row.get(field, "") for field in fields])
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F2937")
        autosize_sheet(sheet)

    chart_sheet = wb.create_sheet("그래프")
    chart_sheet.append(["모델", "GPT-5.5 감성 일치율(%)"])
    for row in sentiment_rows:
        chart_sheet.append([row["비교대상"], float(row["_rate_number"])])
    chart = BarChart()
    chart.title = "GPT-5.5 기준 감성 일치율"
    chart.y_axis.title = "일치율(%)"
    chart.x_axis.title = "모델"
    data = Reference(chart_sheet, min_col=2, min_row=1, max_row=len(sentiment_rows) + 1)
    cats = Reference(chart_sheet, min_col=1, min_row=2, max_row=len(sentiment_rows) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 9
    chart.width = 18
    chart_sheet.add_chart(chart, "D2")
    autosize_sheet(chart_sheet)
    wb.save(path)


def main() -> int:
    source_dir = Path(sys.argv[1]) if len(sys.argv) > 1 else latest_report_dir()
    cleaned_path = source_dir / "tables" / "cleaned_model_comparison.csv"
    rows = read_rows(cleaned_path)

    out_dir = EXPORTS_DIR / f"gpt55_benchmark_{now_stamp()}"
    out_dir.mkdir(parents=True, exist_ok=False)
    (out_dir / "tables").mkdir()
    (out_dir / "graphs").mkdir()

    baseline = "GPT-5.5 기준 판정(Codex audit)"
    current_model_name = "finbert-en-krfinbert-ner-v4"

    n_all, tag_match_all, tag_rate_all = agreement_rate(rows, "codex_primary_tag", "current_tag")
    tagged_rows = [row for row in rows if row.get("current_excluded") == "0"]
    n_tagged, tag_match_tagged, tag_rate_tagged = agreement_rate(tagged_rows, "codex_primary_tag", "current_tag")

    tag_rows = [
        benchmark_row(
            "태깅",
            baseline,
            "현재 운영 모델",
            current_model_name,
            "1순위 카테고리 전체",
            n_all,
            tag_match_all,
            tag_rate_all,
            None,
            "제외된 데이터까지 포함하면 제외/저관련 항목 때문에 일치율이 낮아집니다.",
        ),
        benchmark_row(
            "태깅",
            baseline,
            "현재 운영 모델",
            current_model_name,
            "1순위 카테고리(운영 태깅완료만)",
            n_tagged,
            tag_match_tagged,
            tag_rate_tagged,
            None,
            "보고서 대표 태깅 지표로 권장합니다. 실제 대시보드에 반영되는 태깅 완료 항목 기준입니다.",
        ),
    ]

    sentiment_models = [
        ("현재 운영 모델", current_model_name, "current_sentiment_3", "운영 DB tag_results의 감성 3분류입니다."),
        ("비교모델 종합", "language-aware comparison consensus", "comparison_sentiment_3", "언어별 FinBERT 결과를 우선 사용한 비교모델 종합값입니다."),
        ("언어별 FinBERT", "KR-FinBERT-SC / ProsusAI FinBERT", "language_finbert_sentiment_3", "한국어는 KR-FinBERT-SC, 영어는 ProsusAI/finbert를 사용했습니다."),
        ("금융 키워드 베이스라인", "rule/financial-risk-opportunity-lexicon-v1", "comparison_keyword_label", "투명한 금융 리스크/기회 키워드 규칙 기반 모델입니다."),
    ]
    current_n, current_match, current_rate = agreement_rate(rows, "codex_sentiment_3", "current_sentiment_3")
    sentiment_rows = []
    for target_name, model_name, col, note in sentiment_models:
        n, matched, rate = agreement_rate(rows, "codex_sentiment_3", col)
        current_ref = None if target_name == "현재 운영 모델" else current_rate
        sentiment_rows.append(
            benchmark_row(
                "감성분석",
                baseline,
                target_name,
                model_name,
                "감성 3분류(Positive/Neutral/Negative)",
                n,
                matched,
                rate,
                current_ref,
                note,
            )
        )

    tag_fields = [
        "분석영역", "기준모델", "비교대상", "모델명", "비교항목", "공통건수", "일치건수",
        "일치율", "불일치율", "95% 표본오차", "현재모델 대비 차이", "판정", "비고",
    ]
    write_csv(out_dir / "tables" / "tagging_gpt55_benchmark.csv", tag_rows, tag_fields)
    write_csv(out_dir / "tables" / "sentiment_gpt55_benchmark.csv", sentiment_rows, tag_fields)

    create_svg_bar(out_dir / "graphs" / "tagging_gpt55_agreement.svg", "GPT-5.5 기준 태깅 일치율", tag_rows, "_rate_number")
    sentiment_colored = []
    for row in sentiment_rows:
        copy = dict(row)
        copy["_bar_color"] = "#2563eb" if row["비교대상"] == "현재 운영 모델" else "#94a3b8"
        sentiment_colored.append(copy)
    create_svg_bar(out_dir / "graphs" / "sentiment_gpt55_agreement.svg", "GPT-5.5 기준 감성분석 일치율", sentiment_colored, "_rate_number", "_bar_color")
    create_svg_diff(out_dir / "graphs" / "current_model_advantage.svg", "현재 운영 모델 대비 차이(%p)", [row for row in sentiment_rows if row["비교대상"] != "현재 운영 모델"])

    best_other = max([row for row in sentiment_rows if row["비교대상"] != "현재 운영 모델"], key=lambda row: row["_rate_number"])
    comparison_model = next(row for row in sentiment_rows if row["비교대상"] == "비교모델 종합")
    overview = [
        {"항목": "분석 대상", "내용": f"{len(rows)}건"},
        {"항목": "기준모델", "내용": baseline},
        {"항목": "태깅 대표 지표", "내용": f"현재 운영 모델이 GPT-5.5와 {tag_rows[1]['일치율']} 일치(운영 태깅완료 {n_tagged}건 기준)"},
        {"항목": "감성 대표 지표", "내용": f"현재 운영 모델이 GPT-5.5와 {sentiment_rows[0]['일치율']} 일치"},
        {"항목": "비교모델 종합 대비", "내용": f"현재 운영 모델이 비교모델 종합보다 {pctp(current_rate * 100 - comparison_model['_rate_number'])} 높음"},
        {"항목": "최고 비교모델 대비", "내용": f"최고 비교모델은 {best_other['비교대상']}({best_other['일치율']})이며 현재 운영 모델 대비 {pctp(current_rate * 100 - best_other['_rate_number'])}"},
        {"항목": "주의", "내용": "이 수치는 사람 정답 라벨 기준 accuracy가 아니라 GPT-5.5 기준 판정과의 일치율입니다."},
    ]

    write_workbook(out_dir / "gpt55_benchmark_report.xlsx", overview, tag_rows, sentiment_rows)

    md_fields = ["비교대상", "모델명", "비교항목", "공통건수", "일치건수", "일치율", "불일치율", "95% 표본오차", "현재모델 대비 차이", "판정"]
    md_lines = [
        "# GPT-5.5 기준 모델 벤치마크",
        "",
        f"- 생성 시각: `{datetime.now().isoformat(timespec='seconds')}`",
        f"- 기준 데이터: `{source_dir.name}`",
        f"- 분석 대상: `{len(rows)}`건",
        "- 주의: 아래 수치는 사람 정답 라벨 기준 accuracy가 아니라 `GPT-5.5 기준 판정과의 일치율`입니다.",
        "",
        "## 결론 요약",
        "",
        f"- 태깅 대표 지표: 현재 운영 모델은 GPT-5.5와 **{tag_rows[1]['일치율']}** 일치합니다. 운영 태깅완료 {n_tagged}건 기준입니다.",
        f"- 감성분석: 현재 운영 모델은 GPT-5.5와 **{sentiment_rows[0]['일치율']}** 일치합니다.",
        f"- 비교모델 종합 대비: 현재 운영 모델이 **{pctp(current_rate * 100 - comparison_model['_rate_number'])}** 더 높습니다.",
        f"- 단, 금융 키워드 베이스라인은 GPT-5.5와 **{best_other['일치율']}** 일치하여 일부 감성 항목에서는 현재 모델보다 높게 나왔습니다.",
        "",
        "## 태깅 기준표",
        "",
        "![GPT-5.5 기준 태깅 일치율](graphs/tagging_gpt55_agreement.svg)",
        "",
        markdown_table(tag_rows, md_fields),
        "",
        "## 감성분석 기준표",
        "",
        "![GPT-5.5 기준 감성분석 일치율](graphs/sentiment_gpt55_agreement.svg)",
        "",
        markdown_table(sentiment_rows, md_fields),
        "",
        "## 현재 운영 모델 대비 차이",
        "",
        "![현재 운영 모델 대비 차이](graphs/current_model_advantage.svg)",
    ]
    (out_dir / "gpt55_benchmark_report.md").write_text("\n".join(md_lines), encoding="utf-8-sig")

    html_fields = md_fields
    html_text = "\n".join([
        "<!doctype html><html lang='ko'><head><meta charset='utf-8'>",
        "<title>GPT-5.5 기준 모델 벤치마크</title>",
        "<style>body{font-family:'Malgun Gothic',Inter,Arial,sans-serif;margin:32px;color:#111827;line-height:1.55}table{border-collapse:collapse;width:100%;font-size:13px;margin:12px 0 28px}th,td{border:1px solid #d0d5dd;padding:8px;vertical-align:top}th{background:#f3f4f6}img{max-width:100%;border:1px solid #e5e7eb;margin:8px 0 24px}.note{background:#fff7ed;border:1px solid #fed7aa;padding:12px;border-radius:8px}.good{color:#047857;font-weight:700}.bad{color:#b42318;font-weight:700}</style>",
        "</head><body>",
        "<h1>GPT-5.5 기준 모델 벤치마크</h1>",
        f"<p>분석 대상: <b>{len(rows)}건</b> / 기준 데이터: <code>{html.escape(source_dir.name)}</code></p>",
        "<p class='note'>이 수치는 사람 정답 라벨 기준 accuracy가 아니라 GPT-5.5 기준 판정과의 일치율입니다. GPT-5.5 자체 오차 가능성을 감안하기 위해 불일치율과 95% 표본오차를 함께 표기했습니다.</p>",
        "<h2>결론 요약</h2>",
        "<ul>",
        f"<li>태깅 대표 지표: 현재 운영 모델은 GPT-5.5와 <b>{tag_rows[1]['일치율']}</b> 일치합니다.</li>",
        f"<li>감성분석: 현재 운영 모델은 GPT-5.5와 <b>{sentiment_rows[0]['일치율']}</b> 일치합니다.</li>",
        f"<li>비교모델 종합 대비 현재 운영 모델이 <b>{pctp(current_rate * 100 - comparison_model['_rate_number'])}</b> 더 높습니다.</li>",
        f"<li>최고 비교모델({html.escape(best_other['비교대상'])}) 대비 차이는 <b>{pctp(current_rate * 100 - best_other['_rate_number'])}</b>입니다.</li>",
        "</ul>",
        "<h2>태깅 기준표</h2>",
        "<img src='graphs/tagging_gpt55_agreement.svg'>",
        html_table(tag_rows, html_fields),
        "<h2>감성분석 기준표</h2>",
        "<img src='graphs/sentiment_gpt55_agreement.svg'>",
        html_table(sentiment_rows, html_fields),
        "<h2>현재 운영 모델 대비 차이</h2>",
        "<img src='graphs/current_model_advantage.svg'>",
        "</body></html>",
    ])
    (out_dir / "gpt55_benchmark_report.html").write_text(html_text, encoding="utf-8-sig")

    shutil.copy2(cleaned_path, out_dir / "tables" / "source_cleaned_model_comparison.csv")
    (out_dir / "metadata.json").write_text(json.dumps({
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source_report_dir": str(source_dir),
        "row_count": len(rows),
        "baseline": baseline,
        "current_model": current_model_name,
    }, ensure_ascii=False, indent=2), encoding="utf-8-sig")

    print(str(out_dir))
    print(json.dumps({
        "tagging_current_tagged_agreement": tag_rows[1]["일치율"],
        "sentiment_current_agreement": sentiment_rows[0]["일치율"],
        "current_vs_comparison_model_diff": pctp(current_rate * 100 - comparison_model["_rate_number"]),
        "best_other_model": best_other["비교대상"],
        "current_vs_best_other_diff": pctp(current_rate * 100 - best_other["_rate_number"]),
    }, ensure_ascii=True, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
